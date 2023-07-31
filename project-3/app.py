import openpyxl

def get_user_data():
    data = []
    while True:
        name = input("Enter name (or 'exit' to finish): ")
        if name.lower() == 'exit':
            break
        roll_number = input("Enter roll number: ")
        data.append((name, roll_number))
    return data

def write_to_excel(data):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Student Data"

    # Write the column headers
    sheet.cell(row=1, column=1, value="Name")
    sheet.cell(row=1, column=2, value="Roll Number")

    # Write the data to the sheet
    for idx, (name, roll_number) in enumerate(data, start=2):
        sheet.cell(row=idx, column=1, value=name)
        sheet.cell(row=idx, column=2, value=roll_number)

    # Save the Excel file
    wb.save("student_data.xlsx")

def main():
    print("Enter student data. Type 'exit' for name to finish.")
    data = get_user_data()
    write_to_excel(data)
    print("Data has been successfully written to 'student_data.xlsx'.")

if __name__ == "__main__":
    main()
